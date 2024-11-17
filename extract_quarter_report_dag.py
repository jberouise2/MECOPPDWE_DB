#from my_package.extract_quarter_raw import extract_quarter_raw
from datetime import datetime, timedelta
from airflow import DAG
from airflow.providers.postgres.operators.postgres import PostgresOperator
from airflow.providers.postgres.hooks.postgres import PostgresHook
from airflow.operators.python import PythonOperator
from airflow.sensors.filesystem import FileSensor
from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side

#declaring working directory

#dir = "/opt/airflow/raw_data"

#excel file path
file = "/opt/airflow/raw_data/PPE&DWE_SOUTH_DR_2024.xlsx"

save = "/opt/airflow/data/3rd_QUARTER_ACCOMPLISHMENT_REPORT_2024_test.xlsx"

sheets = ["JUL", "AUG", "SEP"]

def overwrite_sheet(save):

    #excel file path and save path
    save_path = save

    #load the save path file workbook
    wb = load_workbook(save_path)

    # Create a list of sheets to overide in save path
    sheets_to_overide = ["RECEIVED_DATA(DO NOT PRINT)", "TESTED_DATA(DO NOT PRINT)"]

    # Loop through all sheets of save path file and remove the selected sheets
    for sheet_name in sheets_to_overide:
        if sheet_name in wb.sheetnames:
            std = wb[sheet_name]
            wb.remove(std) # Remove the sheet
    wb.save(save_path) #saving empty excel file

def extract_data(file, save, sheets):

    #excel file path and save path
    file_init = file
    save_path = save
   
    # Create a list of sheets to include
    sheets_to_include = sheets

    # Create a list of sheets to overide in save path
    sheets_to_overide = ["RECEIVED_DATA(DO NOT PRINT)", "TESTED_DATA(DO NOT PRINT)"]

    #Reading Each Data Sheet of the Excel File converting it to dataframe
    column_names = ["DATE_RECEIVED","RECEIVED_BY", "SECTOR","COMPANY","QTY","TEST_METHOD","DESCRIPTION", 
                    "DATE_TESTED","TESTED_BY","DETAILS","PASSED","FAILED","RELEASED_DATE","BRAND"]

    #Iterating thru Sheets of file_init
    df_fin = []
        
    for sheet in sheets_to_include:
        df = pd.read_excel(file_init, sheet_name=sheet, skiprows=3, names=column_names)
        df = df[df["DATE_RECEIVED"].notna()]
        df_fin.append(df)

    df_fin = pd.concat(df_fin, ignore_index=False)

    # Convert 'DATE_TESTED' and 'DATE_RECEIVED' to datetime
    df_fin['DATE_TESTED'] = pd.to_datetime(df_fin['DATE_TESTED'], errors='coerce')
    df_fin['DATE_RECEIVED'] = pd.to_datetime(df_fin['DATE_RECEIVED'], errors='coerce')

    #trimming leading and trailing spaces
    for col in ['SECTOR', 'COMPANY', 'DESCRIPTION']:
        df_fin[col] = df_fin[col].str.strip()

    #load workbook
    wb = load_workbook(save_path)


    inquiry_received = df_fin[["DATE_RECEIVED","RECEIVED_BY","SECTOR","COMPANY","QTY","TEST_METHOD","DESCRIPTION"]]
    inquiry_tested = df_fin[["DATE_RECEIVED","RECEIVED_BY","SECTOR","COMPANY","QTY","TEST_METHOD","DESCRIPTION","DATE_TESTED","TESTED_BY","DETAILS","PASSED","FAILED"]]

    with pd.ExcelWriter(save_path, mode='a', engine='openpyxl') as writer:                   
        inquiry_received.to_excel(writer, sheet_name="RECEIVED_DATA(DO NOT PRINT)", index=False)                  
        inquiry_tested.to_excel(writer, sheet_name="TESTED_DATA(DO NOT PRINT)", index=False)

    #load workbook
    wb = load_workbook(save_path)

    # Loop through each sheet
    for sheet in sheets_to_overide:
        ws = wb[sheet]
        ws.sheet_properties.tabColor = "432E54"  # Set the tab color to blue
        # Define border styles
        thin = Side(border_style="thin", color="000000")  # Black color
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Apply borders to all cells
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

    # Save the workbook
    wb.save(save_path)

default_argus= {
    'owner':'jabe',
    'retries': 20,
    'retry_delay': timedelta(minutes=1)
}


with DAG(dag_id='extract_quarter_report_dag',
         default_args=default_argus,
         start_date=datetime(2024, 11, 16),
         end_date=datetime(2024, 11, 17),
         schedule_interval="32 8 * * *",
         catchup=False
         ) as dag:
    
    file_check = FileSensor(
        task_id = 'file_check',
        fs_conn_id = 'fs_default',
        filepath = save,
        mode = 'poke',
        poke_interval = timedelta(minutes=1),
        timeout = timedelta(seconds=300)
    )
    
    overwrite_sheet_task = PythonOperator(
        task_id = 'overwrite_sheet_task',
        python_callable = overwrite_sheet,
        op_kwargs={'save': save}
    )

    extract_data_task = PythonOperator(
        task_id = 'extract_data_task',
        python_callable = extract_data,
        op_kwargs= {
            'file': file,
            'save': save,
            'sheets': sheets
        }
    )

    file_check >> overwrite_sheet_task >> extract_data_task